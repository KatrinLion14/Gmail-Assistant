import pickle
import os
import os.path
import json
import base64
import time
import codecs
import configparser
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from idna import unicode
from openpyxl import Workbook, load_workbook

SCOPES = ['https://www.googleapis.com/auth/gmail.readonly']
name_config = 'config.ini'
section_names = ['urgent_words', 'words_for_labs']


def log_in():
    creds = None
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
            except:
                if os.path.exists('token.pickle'):
                    os.remove(os.path.abspath('token.pickle'))
                flow = InstalledAppFlow.from_client_secrets_file(
                    'credentials.json', SCOPES)
                creds = flow.run_local_server(port=0)
            with open('token.pickle', 'wb') as token:
                pickle.dump(creds, token)
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)
    service = build('gmail', 'v1', credentials=creds)
    return service


def decoding_message(mail):
    data = mail.replace("-", "+").replace("_", "/")
    body = unicode(base64.b64decode(data), 'utf-8')
    return body


def mail_to_json(name, obj):
    f = open(name, 'w', encoding="utf-8")
    f.write(json.dumps(obj, indent=4, ensure_ascii=False))
    f.close()


def mail_to_txt(mail, name):
    for key in mail:
        name.write(unicode(key) + ' : '
                   + unicode(mail[key]) + '\n')


def get_emails(service, query):
    result = service.users().messages().list(userId='me', q=query).execute()
    emails = []
    if result['resultSizeEstimate'] != 0:
        messages = result.get('messages')
        emails = []
        # txts = []
        for msg in messages:
            txt = service.users().messages().get(userId='me', id=msg['id']).execute()
            # txts.append(txt)
            # print(txt)
            try:
                payload = txt['payload']
                headers = payload['headers']
                mail = dict()
                mail['id'] = txt['id']
                subject = ''
                sender = ''
                for d in headers:
                    if d['name'] == 'Subject':
                        subject = d['value']
                    if d['name'] == 'From':
                        sender = d['value']
                parts = payload.get('parts')[0]
                if 'parts' in parts:
                    data = parts['parts'][0]['body']['data']
                    mail['attachment'] = 'true'
                    mail['attachment_names'] = list()
                    mail['attachment_ids'] = list()
                    for i in range(1, len(payload['parts'])):
                        mail['attachment_names'].append(payload['parts'][i]['filename'])
                        mail['attachment_ids'].append(payload['parts'][i]['body']['attachmentId'])
                else:
                    data = parts['body']['data']
                    mail['attachment'] = 'false'
                body = decoding_message(data)
                mail['subject'] = str(subject)
                mail['from'] = str(sender)
                mail['message'] = str(body)
                emails.append(mail)
            except:
                pass
    return emails


def get_attachment(service, messageId, id, prefix, filename):
    att = service.users().messages().attachments().get(userId='me', messageId=messageId, id=id).execute()
    data = att['data']
    file_data = base64.urlsafe_b64decode(data.encode('UTF-8'))
    path = prefix + filename
    with open(path, 'wb') as f:
        f.write(file_data)


def add_keyword_section(section_name, keywords):
    config = configparser.ConfigParser()
    config.read(name_config)
    try:
        config.add_section(section_name)
        section_names.append(section_name)
        i = 0
        for word in keywords:
            config.set(section_name, 'word' + str(i), word)
            i += 1
        with open(name_config, 'w') as configfile:
            config.write(configfile)
    except:
        print('Such section already exists. Do you want to make new one or add to old? 1 or 2')
        answer = int(input())
        if answer == 1:
            print('Give new name')
            section_name = input()
            add_keyword_section(section_name, keywords)
        elif answer == 2:
            i = len(config[section_name])
            for word in keywords:
                config.set(section_name, 'word' + str(i), word)
                i += 1
            with open(name_config, 'w') as configfile:
                config.write(configfile)


def add_word(section_name, word):
    config = configparser.ConfigParser()
    config.read(name_config)
    i = 0
    for option in config[section_name]:
        i += 1
    config.set(section_name, 'word' + str(i), word)
    with open(name_config, 'w') as configfile:
        config.write(configfile)


def delete_section(section_name):
    config = configparser.ConfigParser()
    config.read(name_config)
    config.remove_section(section_name)
    with open(name_config, 'w') as configfile:
        config.write(configfile)


def delete_word(section_name, word):
    config = configparser.ConfigParser()
    config.read(name_config)
    if section_name == '':
        for section in config.sections():
            for item in config[section]:
                if config[section][item] == word:
                    config.remove_option(section, item)
    else:
        for item in config[section_name]:
            if config[section_name][item] == word:
                config.remove_option(section_name, item)


def make_keywords(section_name):
    keywords = []
    config = configparser.ConfigParser()
    config.read(name_config)
    for word in config[section_name]:
        keywords.append(config[section_name][word])
    return keywords


def find_keywords_emails(service, query, section_name, file_name, file_type):
    emails = get_emails(service, query)
    keyword_emails = []
    keywords = make_keywords(section_name)
    for mail in emails:
        for word in keywords:
            if (word in mail['message']) or (word in mail['subject']):
                keyword_emails.append(mail)
    name = str(file_name) + '_messages'
    print_emails(name, keyword_emails, file_type)


def find_urgent_emails(service, file_type):
    urgent_keywords = []
    config = configparser.ConfigParser()
    config.read(name_config)
    for word in config['urgent_words']:
        urgent_keywords.append(config['urgent_words'][word])
    find_keywords_emails(service, '', urgent_keywords, 'urgent', file_type)


def get_all_emails(service, file_type):
    emails = get_emails(service, '')
    print_emails('all_emails', emails, file_type)


def print_emails(name, emails, file_type):
    if file_type == 0:
        file_name = name + '.txt'
        f = codecs.open(file_name, 'w', 'utf-8')
        for mail in emails:
            mail_to_txt(mail, f)
            f.write('===============================\n\n')
    elif file_type == 1:
        file_name = name + '.json'
        mail_to_json(file_name, emails)


def check_labs(emails, labs_number, filename):
    name_addresses_labs = dict()
    wrong_emails = []
    try:    # если файл существует, откроем
        wb = load_workbook(filename + '.xlsx')
        ws = wb.active
        count_rows = ws.max_row
        for i in range(2, count_rows + 1):
            name_addresses_labs[str(ws['A' + str(i)].value).replace('\n', '') + '\r\n'] = i
    except:    # если новый, создадим
        wb = Workbook()
        count_rows = 1
        ws = wb.active
    count_columns = 0
    for i in range(66, 66 + labs_number):
        count_columns += 1
        ws[str(chr(i)) + '1'] = 'lab_' + str(count_columns)
    for mail in emails:
        if mail['attachment'] == 'true':
            if mail['attachment_names'][0].find('lab') > -1:
                name = mail['message']
                if name not in name_addresses_labs:
                    count_rows += 1
                    name_addresses_labs[name] = count_rows
                    ws['A' + str(count_rows)] = name
                for at_name in mail['attachment_names']:
                    lab_name = at_name.split('.')[0]
                    lab_num = lab_name.split('_')[1]
                    ws[chr(65 + int(lab_num)) + str(name_addresses_labs[name])] = '+'
            else:
                keywords = make_keywords('words_for_labs')
                for word in keywords:
                    if (mail['message'].find(word) > -1) or (mail['attachment_names'][0].find(word) > -1):
                        wrong_mail = dict()
                        wrong_mail['body'] = 'Invalid name. Message: ' + mail['message']
                        wrong_mail['attachment names'] = mail['attachment_names']
                        wrong_emails.append(wrong_mail)
    print_emails('wrong_' + str(filename), wrong_emails, 0)
    wb.save(str(filename) + '.xlsx')


def check_course_projects(emails, filename, group_number, file_type):
    wrong_emails = []
    try:  # если файл существует, откроем
        wb = load_workbook(filename + '.xlsx')
        ws = wb.active
        count_rows = ws.max_row
    except:  # если новый, создадим
        wb = Workbook()
        count_rows = 0
        ws = wb.active
    for mail in emails:
        if mail['attachment'] == 'true':
            keywords = make_keywords('words_for_projects')
            for word in keywords:
                if (mail['message'].find(word) > -1) or (mail['subject'].find(word) > -1):
                    if ((mail['message'].find(group_number) > -1) or (mail['subject'].find(group_number) > -1))\
                            and (mail['attachment_names'][0].find(str(file_type)) > -1):
                        count_rows += 1
                        ws['A' + str(count_rows)] = mail['from']
                        ws['B' + str(count_rows)] = mail['attachment_names'][0]
                    else:
                        wrong_mail = dict()
                        wrong_mail['body'] = 'Invalid name. Message: ' + mail['message']
                        wrong_mail['attachment names'] = mail['attachment_names']
                        wrong_emails.append(wrong_mail)
    print_emails('wrong_' + str(filename), wrong_emails, 0)
    wb.save(str(filename) + '.xlsx')


def logout():
    print('Do you want to logout? y/n')
    if input() == 'y':
        if os.path.exists('token.pickle'):
            os.remove(os.path.abspath('token.pickle'))


def main():
    start_time = time.time()
    service = log_in()
    add_word('urgent_words', 'test')
    # print_emails('test', get_emails(service, 'before:2014/02/01'), 0)
    # words = ['test', 'first', 'second']
    # make_new_keyword_section('test', words)
    # print(section_names)
    # get_all_emails(service, 0)
    # get_all_emails(service, 1)
    # check_course_projects(get_emails(service, 'before:2014/02/01'), 'check_projects', '19ПИ-2', 'pdf')
    # find_urgent_emails(service, 0)
    # keyword = []
    # keyword.append('Testing')
    # find_keywords_emails(service, '', keyword, 'testing', 0)
    # logout()
    print("\n\nTIME: ", time.time() - start_time)


if __name__ == '__main__':
    main()
