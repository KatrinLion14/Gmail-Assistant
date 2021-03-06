import pickle
import os
import os.path
import json
import base64
import time
import codecs
import configparser
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from idna import unicode
from openpyxl import Workbook, load_workbook

SCOPES = ['https://www.googleapis.com/auth/gmail.readonly']
name_config = 'config.ini'


def log_in():
    creds = None
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
            except:
                if os.path.exists('token.pickle'):
                    os.remove(os.path.abspath('token.pickle'))
                flow = InstalledAppFlow.from_client_secrets_file(
                    'credentials.json', SCOPES)
                creds = flow.run_local_server(port=0)
            with open('token.pickle', 'wb') as token:
                pickle.dump(creds, token)
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)
    service = build('gmail', 'v1', credentials=creds)
    return service


def decoding_message(mail):
    data = mail.replace("-", "+").replace("_", "/")
    body = unicode(base64.b64decode(data), 'utf-8')
    return body


def mail_to_json(name, obj):
    f = open(name, 'w', encoding="utf-8")
    f.write(json.dumps(obj, indent=4, ensure_ascii=False))
    f.close()


def mail_to_txt(mail, name):
    for key in mail:
        name.write(unicode(key) + ' : '
                   + unicode(mail[key]) + '\n')


def get_emails_for_show(service, query):
    result = service.users().messages().list(userId='me', q=query).execute()
    emails = dict()
    if result['resultSizeEstimate'] != 0:
        messages = result.get('messages')
        for msg in messages:
            txt = service.users().messages().get(userId='me', id=msg['id']).execute()
            try:
                payload = txt['payload']
                headers = payload['headers']
                mail = dict()
                mail['id'] = txt['id']
                subject = ''
                sender = ''
                for d in headers:
                    if d['name'] == 'Subject':
                        subject = d['value']
                    if d['name'] == 'From':
                        sender = d['value']
                parts = payload.get('parts')[0]
                if 'parts' in parts:
                    data = parts['parts'][0]['body']['data']
                    mail['attachment'] = 'true'
                    mail['attachment_names'] = list()
                    mail['attachment_ids'] = list()
                    for i in range(1, len(payload['parts'])):
                        mail['attachment_names'].append(payload['parts'][i]['filename'])
                        mail['attachment_ids'].append(payload['parts'][i]['body']['attachmentId'])
                else:
                    data = parts['body']['data']
                    mail['attachment'] = 'false'
                body = decoding_message(data)
                mail['subject'] = str(subject)
                mail['from'] = str(sender)
                mail['message'] = str(body)
                emails[mail['id']] = mail
            except:
                pass
    return emails


def get_emails(service, query):
    result = service.users().messages().list(userId='me', q=query).execute()
    emails = []
    if result['resultSizeEstimate'] != 0:
        messages = result.get('messages')
        for msg in messages:
            txt = service.users().messages().get(userId='me', id=msg['id']).execute()
            try:
                payload = txt['payload']
                headers = payload['headers']
                mail = dict()
                mail['id'] = txt['id']
                subject = ''
                sender = ''
                for d in headers:
                    if d['name'] == 'Subject':
                        subject = d['value']
                    if d['name'] == 'From':
                        sender = d['value']
                parts = payload.get('parts')[0]
                if 'parts' in parts:
                    data = parts['parts'][0]['body']['data']
                    mail['attachment'] = 'true'
                    mail['attachment_names'] = list()
                    mail['attachment_ids'] = list()
                    for i in range(1, len(payload['parts'])):
                        mail['attachment_names'].append(payload['parts'][i]['filename'])
                        mail['attachment_ids'].append(payload['parts'][i]['body']['attachmentId'])
                else:
                    data = parts['body']['data']
                    mail['attachment'] = 'false'
                body = decoding_message(data)
                mail['subject'] = str(subject)
                mail['from'] = str(sender)
                mail['message'] = str(body)
                emails.append(mail)
            except:
                pass
    return emails


def get_attachment(service, messageId, id, prefix, filename):
    if id == '':
        txt = service.users().messages().get(userId='me', id=messageId).execute()
        try:
            payload = txt['payload']
            parts = payload.get('parts')[0]
            if 'parts' in parts:
                attachment_names = list()
                attachment_ids = list()
                for i in range(1, len(payload['parts'])):
                    attachment_names.append(payload['parts'][i]['filename'])
                    attachment_ids.append(payload['parts'][i]['body']['attachmentId'])
                i = 0
                for attachment in attachment_ids:
                    get_attachment(service, messageId, attachment, prefix, attachment_names[i])
                    i += 1
        except:
            return
    else:
        att = service.users().messages().attachments().get(userId='me', messageId=messageId, id=id).execute()
        data = att['data']
        file_data = base64.urlsafe_b64decode(data.encode('UTF-8'))
        path = prefix + filename
        with open(path, 'wb') as f:
            f.write(file_data)


def add_keyword_section(section_name, keywords):
    config = configparser.ConfigParser()
    config.read(name_config)
    try:
        config.add_section(section_name)
        i = 0
        for word in keywords:
            config.set(section_name, 'word' + str(i), word)
            i += 1
        with open(name_config, 'w') as configfile:
            config.write(configfile)
    except:
        i = len(config[section_name])
        for word in keywords:
            config.set(section_name, 'word' + str(i), word)
            i += 1
        with open(name_config, 'w') as configfile:
            config.write(configfile)


def add_word(section_name, word):
    config = configparser.ConfigParser()
    config.read(name_config)
    i = 0
    for option in config[section_name]:
        i += 1
    config.set(section_name, 'word' + str(i), word)
    with open(name_config, 'w') as configfile:
        config.write(configfile)


def delete_section(section_name):
    config = configparser.ConfigParser()
    config.read(name_config)
    config.remove_section(section_name)
    with open(name_config, 'w') as configfile:
        config.write(configfile)


def delete_word(section_name, word):
    config = configparser.ConfigParser()
    config.read(name_config)
    if section_name == '':
        for section in config.sections():
            for item in config[section]:
                if config[section][item] == word:
                    config.remove_option(section, item)
        with open(name_config, 'w') as configfile:
            config.write(configfile)
    else:
        for item in config[section_name]:
            if config[section_name][item] == word:
                config.remove_option(section_name, item)
        with open(name_config, 'w') as configfile:
            config.write(configfile)


def make_keywords(section_name):
    keywords = []
    config = configparser.ConfigParser()
    config.read(name_config)
    for word in config[str(section_name)]:
        keywords.append(config[section_name][word])
    return keywords


def find_keywords_emails_for_show(service, query, section_name):
    emails = get_emails(service, query)
    keyword_emails = dict()
    keywords = make_keywords(section_name)
    for mail in emails:
        for word in keywords:
            if (word in mail['message']) or (word in mail['subject']):
                keyword_emails[mail['id']] = mail
    return keyword_emails


def find_keywords_emails(service, query, section_name, file_name, file_type, path):
    emails = get_emails(service, query)
    keyword_emails = []
    keywords = make_keywords(section_name)
    for mail in emails:
        for word in keywords:
            if (word in mail['message']) or (word in mail['subject']):
                keyword_emails.append(mail)
    if file_name == '':
        file_name = section_name
    name = str(file_name) + '_messages'
    print_emails(path + name, keyword_emails, file_type)


def find_urgent_emails(service, query, file_type, path, filename):
    find_keywords_emails(service, query, 'urgent_words', filename, file_type, path)


def get_all_emails(service, query, file_type, path, filename):
    emails = get_emails(service, query)
    print_emails(path + filename, emails, file_type)


def print_emails(name, emails, file_type):
    if file_type == 0:
        file_name = name + '.txt'
        f = codecs.open(file_name, 'w', 'utf-8')
        for mail in emails:
            mail_to_txt(mail, f)
            f.write('===============================\n\n')
    elif file_type == 1:
        file_name = name + '.json'
        mail_to_json(file_name, emails)


def check_labs(emails, labs_number, filename, path):
    name_addresses_labs = dict()
    wrong_emails = []
    try:    # если файл существует, откроем
        wb = load_workbook(path + filename + '.xlsx')
        ws = wb.active
        count_rows = ws.max_row
        for i in range(2, count_rows + 1):
            name_addresses_labs[str(ws['A' + str(i)].value).replace('\n', '') + '\r\n'] = i
    except:    # если новый, создадим
        wb = Workbook()
        count_rows = 1
        ws = wb.active
    count_columns = 0
    for i in range(66, 66 + labs_number):
        count_columns += 1
        ws[str(chr(i)) + '1'] = 'lab_' + str(count_columns)
    for mail in emails:
        if mail['attachment'] == 'true':
            if mail['attachment_names'][0].find('lab') > -1:
                name = mail['message']
                if name not in name_addresses_labs:
                    count_rows += 1
                    name_addresses_labs[name] = count_rows
                    ws['A' + str(count_rows)] = name
                for at_name in mail['attachment_names']:
                    lab_name = at_name.split('.')[0]
                    lab_num = lab_name.split('_')[1]
                    ws[chr(65 + int(lab_num)) + str(name_addresses_labs[name])] = '+'
            else:
                keywords = make_keywords('words_for_labs')
                for word in keywords:
                    if (mail['message'].find(word) > -1) or (mail['attachment_names'][0].find(word) > -1):
                        wrong_mail = dict()
                        wrong_mail['body'] = 'Invalid name. Message: ' + mail['message']
                        wrong_mail['attachment names'] = mail['attachment_names']
                        wrong_emails.append(wrong_mail)
    print_emails(path + 'wrong_' + str(filename), wrong_emails, 0)
    wb.save(path + str(filename) + '.xlsx')


def check_course_projects(emails, filename, group_number, file_type, path):
    wrong_emails = []
    try:  # если файл существует, откроем
        wb = load_workbook(path + filename + '.xlsx')
        ws = wb.active
        count_rows = ws.max_row
    except:  # если новый, создадим
        wb = Workbook()
        count_rows = 0
        ws = wb.active
    for mail in emails:
        if mail['attachment'] == 'true':
            keywords = make_keywords('words_for_projects')
            for word in keywords:
                if (mail['message'].find(word) > -1) or (mail['subject'].find(word) > -1):
                    if ((mail['message'].find(group_number) > -1) or (mail['subject'].find(group_number) > -1))\
                            and (mail['attachment_names'][0].find(str(file_type)) > -1):
                        count_rows += 1
                        ws['A' + str(count_rows)] = mail['from']
                        ws['B' + str(count_rows)] = mail['attachment_names'][0]
                    else:
                        wrong_mail = dict()
                        wrong_mail['body'] = 'Invalid name. Message: ' + mail['message']
                        wrong_mail['attachment names'] = mail['attachment_names']
                        wrong_emails.append(wrong_mail)
    print_emails(path + 'wrong_' + str(filename), wrong_emails, 0)
    wb.save(path + str(filename) + '.xlsx')


def log_out():
    if os.path.exists('token.pickle'):
        os.remove(os.path.abspath('token.pickle'))


def main():
    start_time = time.time()
    service = log_in()
    config = configparser.ConfigParser()
    config.read(name_config)
    # print_emails('test', get_emails(service, 'before:2014/02/01'), 0)
    # words = ['test', 'first', 'second']
    # make_new_keyword_section('test', words)
    # print(section_names)
    # get_all_emails(service, 0)
    # get_all_emails(service, 1)
    # check_course_projects(get_emails(service, 'before:2014/02/01'), 'check_projects', '19ПИ-2', 'pdf')
    # find_urgent_emails(service, 0)
    # keyword = []
    # keyword.append('Testing')
    # find_keywords_emails(service, '', keyword, 'testing', 0)
    # logout()
    print("\n\nTIME: ", time.time() - start_time)


if __name__ == '__main__':
    main()
